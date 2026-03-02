"""
check.py
--------
Kiểm tra độ mạnh mật khẩu từ file txt và xuất báo cáo Excel.

Tiêu chí hợp lệ:
  1. Độ dài >= 8 ký tự
  2. Có chữ hoa (A-Z)
  3. Có chữ thường (a-z)
  4. Có chữ số (0-9)
  5. Có ký tự đặc biệt (!@#$%^&*...)

Cách dùng:
  python check.py passwords.txt
  python check.py passwords.txt output.xlsx
  python check.py             ← sẽ hỏi đường dẫn file
"""

import re
import sys
import logging
import traceback
from datetime import datetime
from pathlib import Path

# ── Thư viện bên ngoài (pip install openpyxl) ──────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[LỖI] Thiếu thư viện 'openpyxl'. Chạy: pip install openpyxl")
    sys.exit(1)


# ════════════════════════════════════════════════════════════════════════════
# 1. CẤU HÌNH DEBUG LOGGER
# ════════════════════════════════════════════════════════════════════════════

DEBUG_FILE = "debug.txt"

def _setup_logger() -> logging.Logger:
    """Tạo logger ghi mọi thứ vào debug.txt, chỉ INFO ra console."""
    logger = logging.getLogger("check")
    if logger.handlers:          # tránh thêm handler trùng khi import lại
        return logger
    logger.setLevel(logging.DEBUG)

    fmt = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(funcName)s:%(lineno)d | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # File handler — DEBUG trở lên
    fh = logging.FileHandler(DEBUG_FILE, encoding="utf-8", mode="a")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    # Console handler — INFO trở lên
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)

    logger.addHandler(fh)
    logger.addHandler(ch)

    # Phân cách session trong file log
    with open(DEBUG_FILE, "a", encoding="utf-8") as f:
        f.write(f"\n{'='*70}\n")
        f.write(f"  SESSION: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"{'='*70}\n")

    return logger

log = _setup_logger()


# ══════════════���═════════════════════════════════════════════════════════════
# 2. KIỂM TRA MẬT KHẨU
# ════════════════════════════════════════════════════════════════════════════

MIN_LEN       = 8
RE_UPPER      = re.compile(r"[A-Z]")
RE_LOWER      = re.compile(r"[a-z]")
RE_DIGIT      = re.compile(r"[0-9]")
RE_SPECIAL    = re.compile(r"[!@#$%^&*()\-_=+\[\]{};:'\",.<>?/\\|`~]")


def check_password(account: str, password: str) -> dict:
    """
    Kiểm tra một mật khẩu theo 5 tiêu chí.

    Returns:
        dict {
            "account":  str,
            "password": str,
            "valid":    bool,
            "reasons":  str   (rỗng nếu hợp lệ)
        }
    """
    log.debug(f"Kiểm tra: account='{account}' | len={len(password)}")
    reasons = []

    if len(password) < MIN_LEN:
        reasons.append(f"Độ dài {len(password)} < {MIN_LEN}")

    if not RE_UPPER.search(password):
        reasons.append("Thiếu chữ HOA")

    if not RE_LOWER.search(password):
        reasons.append("Thiếu chữ thường")

    if not RE_DIGIT.search(password):
        reasons.append("Thiếu chữ số")

    if not RE_SPECIAL.search(password):
        reasons.append("Thiếu ký tự đặc biệt")

    is_valid = (len(reasons) == 0)
    log.debug(f"  → {'ĐẠT' if is_valid else 'KHÔNG ĐẠT'}: {reasons}")

    return {
        "account":  account,
        "password": password,
        "valid":    is_valid,
        "reasons":  "; ".join(reasons) if reasons else "",
    }


def read_and_check(filepath: str) -> list[dict]:
    """
    Đọc file txt (định dạng account:password mỗi dòng),
    kiểm tra từng mật khẩu, trả về list kết quả.

    Raises:
        FileNotFoundError, UnicodeDecodeError, ValueError
    """
    log.info(f"Đọc file: {filepath}")
    results = []
    skipped = 0

    try:
        lines = Path(filepath).read_text(encoding="utf-8").splitlines()
    except FileNotFoundError:
        log.error(f"Không tìm thấy file: '{filepath}'")
        raise
    except UnicodeDecodeError as e:
        log.error(f"Lỗi encoding file '{filepath}': {e}", exc_info=True)
        raise

    for i, raw in enumerate(lines, 1):
        line = raw.strip()

        # Bỏ qua dòng trống và comment
        if not line or line.startswith("#"):
            log.debug(f"Dòng {i}: bỏ qua (trống/comment)")
            skipped += 1
            continue

        if ":" not in line:
            log.warning(f"Dòng {i} sai định dạng (thiếu ':'): '{line}'")
            skipped += 1
            continue

        account, _, password = line.partition(":")

        if not account.strip():
            log.warning(f"Dòng {i}: tên tài khoản rỗng")
            skipped += 1
            continue

        results.append(check_password(account.strip(), password))

    log.info(f"Kết quả: {len(results)} tài khoản, {skipped} dòng bỏ qua")

    if not results:
        raise ValueError(f"Không tìm thấy tài khoản hợp lệ nào trong '{filepath}'.")

    return results


# ════════════════════════════════════════════════════════════════════════════
# 3. XUẤT EXCEL
# ════════════════════════════════════════════════════════════════════════════

C_HEADER  = "2F5496"   # Xanh đậm
C_VALID   = "C6EFCE"   # Xanh lá nhạt
C_INVALID = "FFC7CE"   # Đỏ nhạt
C_SUM_H   = "D9E1F2"   # Xanh nhạt tóm tắt


def _border():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _hfont():
    return Font(name="Calibri", bold=True, color="FFFFFF", size=11)

def _dfont(bold=False):
    return Font(name="Calibri", size=10, bold=bold)

def _fill(color):
    return PatternFill("solid", fgColor=color)

def _align(h="left"):
    return Alignment(horizontal=h, vertical="center", wrap_text=True)


def export_excel(results: list[dict], output_path: str) -> str:
    """
    Xuất danh sách kết quả ra file Excel 2 sheet:
      - "Chi tiết": từng tài khoản
      - "Tóm tắt":  thống kê

    Returns:
        Đường dẫn tuyệt đối file Excel vừa tạo.
    """
    log.info(f"Xuất Excel: {output_path}")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Chi tiết ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Chi tiết"

    headers   = ["STT", "Tài khoản", "Mật khẩu", "Trạng thái", "Lý do không đạt"]
    col_widths = [6,      22,          28,          14,            55]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = _hfont()
        c.fill      = _fill(C_HEADER)
        c.alignment = _align("center")
        c.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    valid_count = invalid_count = 0

    for ri, r in enumerate(results, 2):
        is_valid = r["valid"]
        color    = C_VALID if is_valid else C_INVALID
        status   = "✅ Đạt" if is_valid else "❌ Không đạt"
        row_data = [ri - 1, r["account"], r["password"], status,
                    r["reasons"] if r["reasons"] else "Đạt chuẩn"]

        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = _dfont()
            c.fill      = _fill(color)
            c.alignment = _align("center" if ci in (1, 4) else "left")
            c.border    = _border()
        ws.row_dimensions[ri].height = 18

        if is_valid:
            valid_count += 1
        else:
            invalid_count += 1

    ws.freeze_panes = "A2"   # Cố định header khi cuộn

    # ── Sheet 2: Tóm tắt ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Tóm tắt")
    total = len(results)

    sum_headers = ["Chỉ số", "Số lượng", "Tỷ lệ (%)"]
    for ci, h in enumerate(sum_headers, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font      = _hfont()
        c.fill      = _fill(C_HEADER)
        c.alignment = _align("center")
        c.border    = _border()

    rows = [
        ("Tổng số tài khoản",        total,         "100%"),
        ("✅ Số tài khoản ĐẠT",       valid_count,   f"{valid_count/total*100:.1f}%"),
        ("❌ Số tài khoản KHÔNG ĐẠT", invalid_count, f"{invalid_count/total*100:.1f}%"),
    ]
    row_colors = ["FFFFFF", C_VALID, C_INVALID]

    for ri, (label, cnt, pct) in enumerate(rows, 2):
        for ci, val in enumerate([label, cnt, pct], 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font      = _dfont(bold=(ci == 1))
            c.fill      = _fill(row_colors[ri - 2])
            c.alignment = _align("center")
            c.border    = _border()

    for col, w in zip(["A", "B", "C"], [35, 18, 16]):
        ws2.column_dimensions[col].width = w

    wb.save(output_path)
    abs_path = str(Path(output_path).resolve())
    log.info(f"Xuất thành công: {abs_path}")
    return abs_path


# ════════════════════════════════════════════════════════════════════════════
# 4. MAIN
# ════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 55)
    print("  🔐 PASSWORD STRENGTH CHECKER")
    print("=" * 55)

    # ── Lấy đường dẫn file đầu vào ──────────────────────────────────────
    if len(sys.argv) >= 2:
        input_path = sys.argv[1]
    else:
        print("\nNhập đường dẫn file mật khẩu (vd: passwords.txt):")
        input_path = input("  > ").strip()
        if not input_path:
            log.error("Người dùng không nhập đường dẫn.")
            print("❌ Lỗi: chưa nhập đường dẫn file.")
            sys.exit(1)

    # ── Lấy đường dẫn file Excel đầu ra ─────────────────────────────────
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"password_report_{ts}.xlsx"

    log.info(f"input='{input_path}' | output='{output_path}'")

    # ── Đọc & kiểm tra ───────────────────────────────────────────────────
    try:
        print(f"\n🔍 Đọc và kiểm tra: {input_path}")
        results = read_and_check(input_path)
    except FileNotFoundError:
        print(f"❌ Không tìm thấy file '{input_path}'. Xem debug.txt để biết thêm.")
        sys.exit(1)
    except ValueError as e:
        print(f"❌ Dữ liệu không hợp lệ: {e}")
        sys.exit(1)
    except Exception as e:
        log.critical(f"Lỗi không mong đợi:\n{traceback.format_exc()}")
        print(f"❌ Lỗi nghiêm trọng: {e}. Xem debug.txt.")
        sys.exit(1)

    # ── In tóm tắt console ───────────────────────────────────────────────
    total         = len(results)
    valid_count   = sum(1 for r in results if r["valid"])
    invalid_count = total - valid_count

    print("\n" + "─" * 45)
    print(f"  Tổng tài khoản    : {total}")
    print(f"  ✅ Đạt chuẩn      : {valid_count}  ({valid_count/total*100:.1f}%)")
    print(f"  ❌ Không đạt      : {invalid_count}  ({invalid_count/total*100:.1f}%)")
    print("─" * 45)

    # ── Xuất Excel ───────────────────────────────────────────────────────
    try:
        print(f"\n📤 Xuất Excel...")
        abs_path = export_excel(results, output_path)
        print(f"✅ Đã lưu: {abs_path}")
    except PermissionError:
        print(f"❌ Không ghi được '{output_path}' (file đang mở?). Xem debug.txt.")
        sys.exit(1)
    except Exception as e:
        log.critical(f"Lỗi xuất Excel:\n{traceback.format_exc()}")
        print(f"❌ Lỗi xuất Excel: {e}. Xem debug.txt.")
        sys.exit(1)

    print(f"\n📋 Log chi tiết: debug.txt\n")


if __name__ == "__main__":
    main()